"""Regulator-friendly exports from governed Sentinel AML response contracts."""
from __future__ import annotations

import csv
import io
import json
import textwrap
from datetime import datetime, timezone
from typing import Any, Iterable

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer

from agent.models import AgentResponse, AuditEvent, FlaggedEntity


def export_json(data: Any) -> bytes:
    if hasattr(data, "model_dump"):
        data = data.model_dump(mode="json")
    return json.dumps(data, indent=2, default=str).encode("utf-8")


def _flat_entity(entity: FlaggedEntity) -> dict[str, Any]:
    row = entity.model_dump(mode="json")
    row["rule_flags"] = "|".join(row.get("rule_flags") or [])
    row["top_transactions"] = json.dumps(row.get("top_transactions") or [])
    row["risk_contributions"] = json.dumps(row.get("risk_contributions"))
    row["observation_window"] = "|".join(row.get("observation_window") or [])
    return row


def _csv(rows: Iterable[dict[str, Any]], fields: list[str] | None = None) -> bytes:
    materialized = list(rows)
    if not materialized and not fields:
        return b""
    fieldnames = fields or list(materialized[0])
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(materialized)
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def export_entities_csv(entities: list[FlaggedEntity]) -> bytes:
    fields = list(_flat_entity(entities[0])) if entities else [
        "entity_id", "risk_score", "risk_label", "escalation_action",
        "rule_flags", "explanation", "sar_draft",
    ]
    return _csv((_flat_entity(entity) for entity in entities), fields)


def export_entities_xlsx(response: AgentResponse) -> bytes:
    workbook = Workbook()
    entities_sheet = workbook.active
    entities_sheet.title = "Flagged entities"
    entity_rows = [_flat_entity(entity) for entity in response.top_entities]
    fields = list(entity_rows[0]) if entity_rows else [
        "entity_id", "risk_score", "risk_label", "escalation_action",
    ]
    entities_sheet.append(fields)
    for row in entity_rows:
        entities_sheet.append([row.get(field) for field in fields])
    entities_sheet.freeze_panes = "A2"
    entities_sheet.auto_filter.ref = entities_sheet.dimensions

    trace_sheet = workbook.create_sheet("Execution trace")
    trace_fields = ["tool", "status", "duration_ms", "reason"]
    trace_sheet.append(trace_fields)
    for step in response.execution_trace:
        payload = step.model_dump()
        trace_sheet.append([payload[field] for field in trace_fields])
    trace_sheet.freeze_panes = "A2"

    evidence_sheet = workbook.create_sheet("Transaction evidence")
    evidence_fields = [
        "entity_id", "txn_id", "timestamp", "amount", "payment_format",
        "to_account", "triggered_rules",
    ]
    evidence_sheet.append(evidence_fields)
    for entity in response.top_entities:
        for transaction in entity.top_transactions:
            payload = transaction.model_dump()
            evidence_sheet.append(
                [
                    entity.entity_id,
                    payload["txn_id"],
                    payload["timestamp"],
                    payload["amount"],
                    payload["payment_format"],
                    payload["to_account"],
                    "|".join(payload["triggered_rules"]),
                ]
            )
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_trace_csv(response: AgentResponse) -> bytes:
    return _csv(
        (step.model_dump(mode="json") for step in response.execution_trace),
        ["tool", "status", "duration_ms", "reason"],
    )


def export_audit_csv(events: list[AuditEvent]) -> bytes:
    fields = [
        "event_id", "dataset_id", "event_type", "actor", "investigation_id",
        "alert_id", "payload", "risk_policy_version", "model_version",
        "dataset_snapshot", "created_at",
    ]
    rows = []
    for event in events:
        row = event.model_dump(mode="json")
        row["payload"] = json.dumps(row["payload"], separators=(",", ":"))
        rows.append(row)
    return _csv(rows, fields)


def export_sar_txt(entity: FlaggedEntity) -> bytes:
    start, end = entity.observation_window or ("Not established", "Not established")
    text = f"""SUSPICIOUS ACTIVITY REPORT (DRAFT)
Generated at: {datetime.now(timezone.utc).isoformat()}

Subject account: {entity.entity_id}
Suspicious period: {start} to {end}
Typology: {entity.saml_d_typology or "Unclassified"}
Risk score: {entity.risk_score:.2f} ({entity.risk_label.upper()})
Recommended action: {entity.escalation_action}

DESCRIPTION OF SUSPICIOUS ACTIVITY
{entity.sar_draft or entity.explanation or "No narrative was generated."}

RULE SIGNALS
{", ".join(entity.rule_flags) if entity.rule_flags else "No rule flags recorded."}

EVIDENCE CITATION
{entity.citation or "No citation recorded."}

REVIEWER NOTICE
DRAFT ONLY. A qualified human reviewer must validate evidence, complete all
institution-specific fields, and approve any regulatory filing.
"""
    return text.encode("utf-8")


def export_investigation_md(
    response: AgentResponse,
    dataset_name: str | None = None,
) -> bytes:
    lines = [
        "# Sentinel AML Investigation Report",
        "",
        f"- Investigation ID: `{response.investigation_id or 'not persisted'}`",
        f"- Dataset: {dataset_name or response.dataset_name or 'Active dataset'}",
        f"- Dataset ID: `{response.dataset_id or 'active'}`",
        f"- Query: {response.query}",
        f"- Intent: {response.intent.intent}",
        f"- Pattern: {response.intent.pattern_type or 'general'}",
        "",
        "## Execution plan",
        "",
        "| Tool | Status | Duration (ms) | Reason |",
        "|---|---|---:|---|",
    ]
    lines.extend(
        f"| {step.tool} | {step.status} | {step.duration_ms:.3f} | "
        f"{step.reason.replace('|', '/')} |"
        for step in response.execution_trace
    )
    stats = response.summary_stats
    lines += [
        "",
        "## Risk summary",
        "",
        f"- Rows analyzed: {stats.total_analyzed:,}",
        f"- Flagged entities: {stats.flagged:,}",
        f"- High-risk entities: {stats.high_risk:,}",
        "",
        "## Flagged entities",
        "",
    ]
    for entity in response.top_entities:
        lines += [
            f"### {entity.entity_id} — {entity.risk_label.upper()} "
            f"({entity.risk_score:.2f})",
            "",
            f"- Recommended action: {entity.escalation_action}",
            f"- Typology: {entity.saml_d_typology or 'Unclassified'}",
            f"- Rules: {', '.join(entity.rule_flags) or 'None'}",
            f"- Explanation: {entity.explanation or 'Not available'}",
            f"- SAR draft: {entity.sar_draft or 'Not available'}",
            "",
        ]
    lines += [
        "## Governance signature",
        "",
        "- Risk policy: sentinel-risk-policy-1.0.0",
        "- Model: sentinel-saml-iforest-v1",
        "- Human review is required before reporting or filing.",
    ]
    return ("\n".join(lines) + "\n").encode("utf-8")


def export_model_card_md(card: dict[str, Any]) -> bytes:
    lines = ["# Sentinel AML Model Card", ""]
    for key, value in card.items():
        if isinstance(value, (dict, list)):
            lines += [
                f"## {key.replace('_', ' ').title()}",
                "",
                "```json",
                json.dumps(value, indent=2, default=str),
                "```",
                "",
            ]
        else:
            lines.append(f"- **{key.replace('_', ' ').title()}:** {value}")
    return ("\n".join(lines) + "\n").encode("utf-8")


def export_pdf(title: str, sections: list[tuple[str, str]]) -> bytes:
    output = io.BytesIO()
    styles = getSampleStyleSheet()
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=18 * mm,
        leftMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=title,
        author="Sentinel AML",
    )
    story = [Paragraph(title, styles["Title"]), Spacer(1, 8)]
    for heading, body in sections:
        if heading:
            story.extend([Paragraph(heading, styles["Heading2"]), Spacer(1, 4)])
        paragraphs = body.splitlines() or [""]
        for paragraph in paragraphs:
            safe = (
                paragraph.replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )
            for chunk in textwrap.wrap(safe, 120) or [""]:
                story.append(Paragraph(chunk or " ", styles["BodyText"]))
            story.append(Spacer(1, 4))
    document.build(story)
    return output.getvalue()


def export_sar_pdf(entity: FlaggedEntity) -> bytes:
    text = export_sar_txt(entity).decode("utf-8")
    return export_pdf(
        f"SAR Draft — {entity.entity_id}",
        [("Draft narrative and evidence", text)],
    )


def export_investigation_pdf(response: AgentResponse) -> bytes:
    markdown = export_investigation_md(response).decode("utf-8")
    return export_pdf(
        f"Investigation {response.investigation_id or ''}".strip(),
        [("Governed investigation record", markdown)],
    )


def export_model_card_pdf(card: dict[str, Any]) -> bytes:
    return export_pdf(
        "Sentinel AML Model Card",
        [("Validated model metadata", json.dumps(card, indent=2, default=str))],
    )
